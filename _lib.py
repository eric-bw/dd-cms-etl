from openpyxl import load_workbook, Workbook
import datetime
import base64
import requests
import csv
import os
import shutil
import zipfile
import re

def str2bool(v):
    return v.lower() in ("yes", "true", "t", "1")

class Transfer:
    def __init__(self):
        self.a = None
        self.b = None
        self.a_id = None
        self.b_id = None

def create_pages(sf):
    wb = load_workbook('input/site.xlsx')
    for sheet_name in wb.sheetnames:
        if sheet_name == 'pages':
            sheet = wb[sheet_name]
            for i, row in enumerate(sheet.rows):
                if i > 0:
                    q = 'select id from CMS_Page__c where uri__c = \'%s\''%(row[3].value,)
                    rs = sf.query(q)['records']
                    if rs:
                        sheet[row[0].coordinate] = rs[0]['Id']
                    else:
                        record = sf.CMS_Page__c.create({'Name':row[1].value,
                                                        'breadcrumb__c':row[2].value,
                                                        'Story__c': row[3].value,
                                                        'uri__c': row[4].value})
                        sheet[row[0].coordinate] = record['id']
    wb.save('input/site.xlsx')

def output(sf, filepath, args) :
    wb = Workbook()
    wb.remove_sheet(wb.active)
    filter=set()

    if args.mega or args.pages:
        print('exporting filtered content for ' + args.username.split('@')[1])
        filepath = 'filtered_' + filepath
        print('filtering by mega menus:', args.mega, ', pages', args.pages)
    else:
        print('exporting content for ' + args.username.split('@')[1])

    add_sheet(sf, wb, ['Id'], 'CMS_Mega_Menu__c', 'Mega Menu', filter, args)
    print('.', end='')
    add_sheet(sf, wb, ['Id'], 'CMS_Page__c', 'Pages', filter, args)
    print('.', end='')
    add_sheet(sf, wb, ['Id'], 'CMS_Content__c', 'Contents', filter, args)
    print('.', end='')
    add_sheet(sf, wb, ['Id'],'CMS_Asset__c', 'Assets', filter, args)
    print()
    print('creating content pack')
    add_sheet_content(sf, wb, ['Id','ContentSize','Checksum'], 'ContentVersion', 'Files', filepath, filter, args)
    wb.save(filepath + '.xlsx')

    zipf = zipfile.ZipFile(filepath + '.zip', 'w')
    zipf.write(filepath + '.xlsx')
    for f in os.listdir('./content'):
        zipf.write('./content/' + f)
    os.remove(filepath + '.xlsx')
    shutil.rmtree('./content')
    print()
    print('contentpak:', filepath + '.zip')
    print('done')

def should_filter(row, object_name, filter_state, args):
    slug = ''
    if 'Slug__c' in row:
        slug = row['Slug__c']

    if object_name == 'CMS_Mega_Menu__c':
        if args.mega and slug in args.mega:
            filter_state.add(slug)
            return False
        if not args.mega and not args.pages :
            return False
        return True
    elif object_name == 'CMS_Page__c':
        if args.pages and slug in args.pages:
            filter_state.add(slug)
            return False
        if not args.mega and not args.pages :
            return False
        return True
    elif object_name == 'CMS_Content__c':
        if not filter_state:
            return False
        should_filter = True
        if row['CMS_Page__r'] and row['CMS_Page__r']['Slug__c'] in filter_state:
            should_filter = False
        elif row['Collection__r'] and row['Collection__r']['CMS_Page__r'] and row['Collection__r']['CMS_Page__r']['Slug__c'] in filter_state:
            should_filter = False
        elif row['CMS_Mega_Menu__r'] and row['CMS_Mega_Menu__r']['Slug__c'] in filter_state:
            should_filter = False
        elif row['Collection__r'] and row['Collection__r']['CMS_Mega_Menu__r']and row['Collection__r']['CMS_Mega_Menu__r']['Slug__c'] in filter_state:
            should_filter = False

        if not should_filter:
            filter_state.add(row['Id'])
        return should_filter
    elif object_name == 'CMS_Asset__c':
        if not filter_state:
            return False
        if row['CMS_Content__c'] in filter_state:
            filter_state.add(row['ContentDocument__c'])
            return False
        return True
    elif object_name == 'ContentVersion':
        if not filter_state:
            return False
        if row['ContentDocumentId'] in filter_state:
            return False
        return True
    else:
        pass

def add_sheet(sf, wb, fields,  object_name, sheet_name, filter_state, args):
    meta = sf.__getattr__(object_name).describe()
    recordtypes = {}
    for r in meta['recordTypeInfos']:
        recordtypes[r['recordTypeId']] = r['name']
    if object_name == 'CMS_Content__c':
        records = sf.query('select Collection__r.CMS_Page__r.Slug__c, CMS_Page__r.Slug__c, Collection__r.CMS_Mega_Menu__r.Slug__c, CMS_Mega_Menu__r.Slug__c, ' + ','.join(fields) +','+ ','.join([x['name'] for x in meta['fields'] if x['createable']]) + ' from ' + object_name + ' order by collection__c desc')['records']
    else:
        records = sf.query('select ' + ','.join(fields) +','+ ','.join([x['name'] for x in meta['fields'] if x['createable']]) + ' from ' + object_name + ' order by createddate')['records']
    wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]

    for i, row in enumerate(records):
        if i == 0:
            th = [x for x in row.keys() if x not in ['attributes','OwnerId','CMS_Page__r','Collection__r','CMS_Mega_Menu__r']]
            if 'RecordTypeId' in row:
                th = th + ['RecordTypeName']
            sheet.append(th)

        if should_filter(row, object_name, filter_state, args):
            continue

        if 'CMS_Page__r' in row: del row['CMS_Page__r']
        if 'Collection__r' in row: del row['Collection__r']
        if 'CMS_Mega_Menu__r' in row: del row['CMS_Mega_Menu__r']

        tr = [y for x,y in row.items() if x not in ['attributes','OwnerId']]
        if object_name == 'CMS_Content__c' and 'RecordTypeId' in row:
            if row['RecordTypeId']:
                tr = tr + [recordtypes[row['RecordTypeId']]]
            else:
                raise Exception('missing recordtype for content', row['Id'], 'update and try again')

        sheet.append(tr)

def add_sheet_content(sf, wb, fields,  object_name, sheet_name, filepath, filter_state, args):
    meta = sf.__getattr__(object_name).describe()
    records = sf.query('select ' + ','.join(fields) +','+ ','.join([x['name'] for x in meta['fields'] if x['createable']]) +
                       ' from ' + object_name +
                       ' where ContentDocumentId in (\'' + '\',\''.join(get_file_list(sf)) +'\') ' +
                       ' order by createddate')['records']
    wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]
    if os.path.exists('./content'):
        shutil.rmtree('./content')
    os.mkdir('./content')
    for i, row in enumerate(records):
        if i == 0:
            th = [x for x in row.keys() if x not in ['attributes','OwnerId']]
            sheet.append(th)
        tr = [y for x,y in row.items() if x not in ['attributes','OwnerId']]

        if should_filter(row, object_name, filter_state, args):
            continue
        print('.', end='')
        url = "https://%s%s" % (sf.sf_instance, row['VersionData'])
        response = requests.get(url, headers={"Authorization": "OAuth " + sf.session_id, "Content-Type": "application/octet-stream"})
        if response.ok:
            byte_string = str(base64.b64encode(response.content))
            byte_string = byte_string[2:-1]
            with open('./content' + '/' + row['ContentDocumentId'],'w') as f:
                f.write(byte_string)
                f.close()
        else:
            print('file err: ' + response)
            open('./content' + '/' + row['ContentDocumentId'],'w')


        sheet.append(tr)

def get_data(sheet, key='Id'):
    rs = {}
    field_map = {}
    for i, row in enumerate(sheet.rows):
        if i == 0:
            for n, f in enumerate(row):
                field_map[n] = f.value
        else:
            data = {}
            id = None
            for n, f in enumerate(row):
                header = field_map[n]
                if header == key:
                    id = f.value
                    continue
                data[header] = f.value
            t = Transfer()
            t.a = data
            t.a_id = id
            rs[id] = t
    return rs

def create_record(meta, row):
    rs = {}
    for base_field in row:
        found = False
        if base_field == 'RecordTypeName':
            rs[base_field] = row[base_field]
            continue
        for target_field in meta['fields']:
            if base_field == target_field['name']:
                rs[base_field] = row[base_field]
                found = True
                break
        if not found:
            # print(base_field, 'not found')
            pass

    return rs

def transfer_pages(map, wb, wb_out, sf, read_only=False):
    sheet_name = 'Pages'
    sheet = wb[sheet_name]
    data = get_data(sheet, 'Id')
    meta = sf.CMS_Page__c.describe()


    for row in data.values():
        note = ''
        record = create_record(meta, row.a.copy())
        exists = sf.query('select id, Name from CMS_Page__c where slug__c = \'%s\''%(row.a['Slug__c']))['records']
        if exists:
            row.b_id = exists[0]['Id']
            row.b = exists[0]
            if not read_only: sf.CMS_Page__c.update(row.b_id, record)
            note += 'updated record'
            print('.', end='')
        else:
            if not read_only:
                try:
                    rs = sf.CMS_Page__c.create(row.a)
                    row.b_id = rs['id']
                    row.b = row.a
                    note += 'created record'
                    print('+', end='')
                except Exception as e:
                    print(e)
            else:
                raise Exception('record does not exist' + record['Name'])
        record['Id'] = row.b_id
        wb_out.writerow([row.a_id,row.b_id, note])

    map.update(data)
    print()

def get_recordtype(meta, recordtype_name):
    for rt in meta['recordTypeInfos']:
        if rt['name'] == recordtype_name:
            return rt
    return None

def transfer_mega(map, wb, wb_out, sf, read_only=False):
    sheet_name = 'Mega Menu'
    sheet = wb[sheet_name]
    data = get_data(sheet)
    meta = sf.CMS_Mega_Menu__c.describe()
    for row in data.values():
        note = ''
        record = create_record(meta, row.a.copy())

        exists = sf.query('select id, Name from CMS_Mega_Menu__c where slug__c = \'%s\''%(row.a['Slug__c']))['records']


        if exists:
            row.b_id = exists[0]['Id']
            row.b = exists[0]
            if not read_only: sf.CMS_Mega_Menu__c.update(row.b_id, record)
            note += 'updated record'
            print('.', end='')
        else:
            if not read_only:
                rs = sf.CMS_Mega_Menu__c.create(record)
                row.b_id = rs['id']
                row.b = row.a
                note += 'created record'
                print('+', end='')
            else:
                raise Exception('record does not exist' + record['Name'])
        record['Id'] = row.b_id
        wb_out.writerow([row.a_id,row.b_id, note])
    print()
    map.update(data)

def transfer_content(map, wb, wb_out, sf, read_only=False):
    sheet_name = 'Contents'
    sheet = wb[sheet_name]
    data = get_data(sheet)
    meta = sf.CMS_Content__c.describe()

    for row in data.values():
        note = ''
        record = create_record(meta, row.a.copy())

        if record['Collection__c'] in map:
            record['Collection__c'] = map[record['Collection__c']].b_id

        if record['CMS_Page__c'] in map:
            record['CMS_Page__c'] = map[record['CMS_Page__c']].b_id

        if record['CMS_Mega_Menu__c'] in map:
            record['CMS_Mega_Menu__c'] = map[record['CMS_Mega_Menu__c']].b_id

        recordtypes={}
        meta = sf.CMS_Content__c.describe()
        for r in meta['recordTypeInfos']:
            recordtypes[r['name']] = r['recordTypeId']

        if record['RecordTypeName'] in recordtypes:
            record['RecordTypeId'] = recordtypes[record['RecordTypeName']]
            record.pop('RecordTypeName')

        exists = sf.query('select id, Name from CMS_Content__c where slug__c = \'%s\''%(row.a['Slug__c']))['records']


        if exists:
            row.b_id = exists[0]['Id']
            row.b = exists[0]
            if not read_only: sf.CMS_Content__c.update(row.b_id, record)
            note += 'updated record'
            print('.', end='')
        else:
            if not read_only:
                rs = sf.CMS_Content__c.create(record)
                row.b_id = rs['id']
                row.b = row.a
                note += 'created record'
                print('+', end='')
            else:
                raise Exception('record does not exist' + record['Name'])
        record['Id'] = row.b_id
        wb_out.writerow([row.a_id,row.b_id, note])
        map.update(data) #need to update with every creation since the its a nested structure
    print()

def transfer_files(map, wb, wb_out, sf, archive, read_only=False):
    sheet_name = 'Files'
    sheet = wb[sheet_name]
    data = get_data(sheet, 'ContentDocumentId')

    meta = sf.ContentVersion.describe()
    for document_id, row in data.items():
        note = ''
        record = create_record(meta, row.a.copy())
        exists = sf.query('select id, ContentDocumentId, Description  from ContentVersion where PathOnClient = \'%s\' and Checksum = \'%s\''%(row.a['PathOnClient'],row.a['Checksum']))['records']
        if exists:
            row.b_id = exists[0]['Id']
            row.b = exists[0]
            row.b['Description'] = 'cms-asset'

            if not read_only: sf.ContentDocument.update(row.b['ContentDocumentId'], {'Description':'cms-asset'})
            note += 'updated record'
            print('.', end='')
        else:
            if not read_only:
                b64encoded = archive.read('content/' + document_id).decode('utf-8')
                row.b = {'title' : row.a['Title'],'PathOnClient' : row.a['PathOnClient'],'VersionData' : b64encoded, 'Description': 'cms-asset'}
                content = sf.ContentVersion.create(row.b)
                row.b_id = content['id']
                row.b = sf.query('select id, Title, VersionData, PathOnClient, ContentDocumentId from ContentVersion where id = \'%s\''%(row.b_id))['records'][0]
                note += 'created record'
                print('+', end='')
            else:
                raise Exception('record does not exist' + record['Title'])
        record['Id'] = row.b_id

        library = sf.query('select id from ContentWorkspace where DeveloperName = \'sfdc_asset_company_assets\'')['records']
        if not library:
            raise Exception('Asset Library not found (sfdc_asset_company_assets)')

        share = sf.query('select id from ContentDocumentLink where ContentDocumentId = \'%s\' and LinkedEntityId = \'%s\''%(row.b['ContentDocumentId'], library[0]['Id']))['records']
        if not share:
            rs = sf.ContentDocumentLink.create({'ContentDocumentId': row.b['ContentDocumentId'], 'LinkedEntityId': library[0]['Id'], 'ShareType':'I','Visibility':'AllUsers'})
            wb_out.writerow(['',rs['id'], 'created content share'])

        wb_out.writerow([row.a_id,row.b_id, note])
    print()
    map.update(data)

def get_file_list(sf):
    rs = []
    for row in sf.query('select ContentDocument__c from CMS_Asset__c')['records']:
        rs.append(row['ContentDocument__c'])
    return rs

def transfer_assets(map, wb, wb_out, sf):
    sheet_name = 'Assets'
    sheet = wb[sheet_name]
    data = get_data(sheet)

    meta = sf.CMS_Asset__c.describe()

    for row in data.values():
        note = ''
        record = create_record(meta, row.a.copy())

        if record['CMS_Content__c'] in map:
            record['CMS_Content__c'] = map[record['CMS_Content__c']].b_id
        else:
            note += 'record missing content %s;'%(record['CMS_Content__c'])
            print(note)

        if 'Type__c' in record:
            record['Asset_Type__c'] = record['Type__c']
            record.pop('Type__c')

        if 'type__c' in record:
            record['Asset_Type__c'] = record['type__c']
            record.pop('type__c')


        if not record['Asset_Type__c']:
            record['Asset_Type__c'] = 'MAIN_IMAGE'

        if row.a['ContentDocument__c'] in map:
            record['ContentDocument__c'] = map[row.a['ContentDocument__c']].b['ContentDocumentId']
            record['ContentVersion__c'] = map[row.a['ContentDocument__c']].b['Id']
            record['Name'] = map[row.a['ContentDocument__c']].a['PathOnClient']
        else:
            note += 'record %s missing version;'%(record['ContentVersion__c'])

        exists = sf.query('select id, Name from CMS_Asset__c where slug__c = \'%s\''%(slugify(record['Asset_Type__c'], record['CMS_Content__c'])))['records']
        if exists:
            row.b_id = exists[0]['Id']
            row.b = exists[0]
            try:
                rs = sf.CMS_Asset__c.update(row.b_id, record)
                note += 'updated record'
                print('.', end='')
            except Exception as e:
                note += 'updated failed -- %s;'%(e.content[0]['message'])
        else:
            rs = sf.CMS_Asset__c.create(record)
            row.b_id = rs['id']
            row.b = row.a
            note += 'created record'
            print('+', end='')
        row.a['Id'] = row.b_id
        wb_out.writerow([row.a_id,row.b_id, note])
    map.update(data)
    print()

def slugify(*args):
    return re.sub('[^0-9a-zA-Z]+', '-', (' '.join(args)).lower())

def transfer(filepath, sf, wb_out, args, read_only = False):
    archive = zipfile.ZipFile(filepath, 'r')
    excel_file = None
    for x in archive.namelist():
        if x.endswith('.xlsx'):
            excel_file = archive.open(x)
            break

    if excel_file is None:
        raise Exception('file not found')

    wb = load_workbook(excel_file)
    map={}
    print('transferring files')
    transfer_files(map, wb, wb_out, sf, archive, read_only)
    print('transferring pages')
    transfer_pages(map, wb, wb_out,  sf, read_only)
    print('transferring mega menus')
    transfer_mega(map, wb, wb_out, sf, read_only)
    print('transferring content')
    transfer_content(map, wb, wb_out, sf, read_only)
    print('transferring assets')
    transfer_assets(map, wb, wb_out, sf)
    print('done')

def clear_content(sf, objects):
    print('clearing existing data')
    for o in ['ContentDocument','CMS_Mega_Menu__c','CMS_Page__c','CMS_Content__c','CMS_Asset__c']:
        if objects and o not in objects:
            continue
        try:
            print('deleting ' + o)
            records = []
            if o == 'ContentDocument':
                scope = []
                for x in sf.query('select ContentDocument__c from CMS_Asset__c')['records']:
                    scope.append(x['ContentDocument__c'])
                for x in sf.query('select id from ' + o + ' where id in (\'%s\')'%("','".join(scope)))['records']:
                    records.append({'Id':x['Id']})
            else:
                for x in sf.query('select id from ' + o)['records']:
                    records.append({'Id': x['Id']})
            if records:
                sf.bulk.__getattr__(o).delete(records)
            else:
                print('no records found')
        except Exception as e:
            print('err: ' + str(e) )
    print('done')