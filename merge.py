import sys
if sys.version_info < (3, 0):
    sys.stdout.write("Sorry, this software requires Python 3(.7), not Python " + str(sys.version_info.major) + '.' +  str(sys.version_info.minor) + "\n")
    sys.exit(1)

from openpyxl import load_workbook, Workbook
import datetime
import _lib
import zipfile

import argparse
import sys

import re
from simple_salesforce import Salesforce
import difflib

parser = argparse.ArgumentParser(description='transfer content from a specifically designed excel document into a new org')

parser.add_argument('-u', '--username',
                    help=' Username',
                    required=True,
                    default='.')

parser.add_argument('-p', '--password',
                    help=' Password',
                    required=True,
                    default='.')

parser.add_argument('-t', '--token',
                    help=' Token',
                    required=False,
                    default='')

parser.add_argument('-s', '--sandbox',
                    type=_lib.str2bool,
                    help=' is sandbox',
                    required=True,
                    default=True)

parser.add_argument('-p1', '--primary',
                    help='ContentPack 1',
                    required=True,
                    default='.')

parser.add_argument('-p2', '--secondary',
                    help=' Content Pack 2',
                    required=True,
                    default='.')

parser.add_argument('-m', '--merge',
                    type=_lib.str2bool,
                    help=' execute merge',
                    required=False,
                    default=False)

parser.add_argument('-i', '--interactive',
                    type=_lib.str2bool,
                    help=' Interactively update env',
                    required=False,
                    default=True)



args = parser.parse_args(sys.argv[1:])

def get_x(path):
    archive = zipfile.ZipFile(path, 'r')
    excel_file = None
    for x in archive.namelist():
        if x.endswith('.xlsx'):
            excel_file = archive.open(x)
            break
    return excel_file



def compare_tab(tab, obj_name, field, f1, f2, args, dev, pages={}, content={}):
    sheet1 = f1[tab]
    sheet2 = f2[tab]
    data = {}

    for i, row in enumerate(sheet1.rows):
        if i == 0:
            columns = {}
            for cell in row:
                columns[cell.value] = cell.col_idx - 1
                columns[cell.col_idx - 1] = cell.value
        if i > 0:
            id = row[columns[field]].value
            record = {columns[i]:x.value for i,x in enumerate(row)}
            data[id] = {'a':record}
    for i, row in enumerate(sheet2.rows):
        if i == 0:
            columns = {}
            for cell in row:
                columns[cell.value] = cell.col_idx - 1
                columns[cell.col_idx - 1] = cell.value
        if i > 0:
            id = row[columns[field]].value
            record = {columns[i]:x.value for i,x in enumerate(row)}
            if id in data:
                data[id]['b'] = record
            else:
                data[id] = {'b':record}
    compare(tab, obj_name, data, args, dev, pages)
    #print('comparing tab', tab, 'complete')
    return data

def get_page(id, pages):
    for p in pages.values():
        if 'a' in p and 'b' in p and p['b']['Id'] == id:
            return p['a']
    return None

def compare(tab, obj_name, data, args, dev, pages={}, content={}):
    for row in data.values():
        if 'a' not in row:
            record = row['b'].copy()
            id = record.pop('Id')
            if 'CMS_Page__c' in record:
                if record['CMS_Page__c'] is None:
                    record.pop('CMS_Page__c')
                else:
                    page = get_page(record['CMS_Page__c'], pages)
                    record['CMS_Page__c'] = page['Id']
            if 'Collection__c' in record:
                record.pop('Collection__c')
            if 'RecordTypeName' in record:
                _lib.set_recordtype(dev, record)
            rs = dev.__getattr__(obj_name).create(record)
            print(rs)
            continue
        if 'b' not in row:
            # print(tab, '(' + row['a']['Id'] + ')', row['a']['Slug__c'])
            # print('not in ', args.secondary)
            continue
        for field, value in row['a'].items():

            if field == 'Id' or field == 'Collection__c' or field == 'RecordTypeId' or field == 'CMS_Page__c' or field == 'CMS_Mega_Menu__c': continue
            if 'readcrumb__c' in field: continue #temporary
            if field not in row['b']: continue
            new_value = row['b'][field]
            if re.sub("\s*", "", repr(value).replace('\\r\\n','\\n')) != re.sub("\s*", "", repr(new_value).replace('\\r\\n','\\n')):
                print(tab, '(' + row['a']['Id'] + ')')
                print(args.primary.ljust(55), field, '=',repr(value).replace('\\r\\n','\\n'))
                print(args.secondary.ljust(55), field,'=', repr(new_value))
                for x in difflib.Differ().compare(repr(value).replace('\\r\\n','\\n'), repr(new_value).replace('\\r\\n','\\n')):
                    if x.startswith('+') or x.startswith('-'):
                        print(x, end=', ')
                print('------')
                do_update = False
                # if args.merge:
                #     do_update = True
                #     if args.interactive:
                #         rs = input('update change?')
                #         if rs == 'y':
                #             do_update = True
                #         else:
                #             do_update = False
                #             print('skipping')
                # if do_update:
                #     rs = dev.__getattr__(obj_name).update(row['a']['Id'], { field:new_value})
                #     print(rs)

def get_content(map, field, sheet1, sheet2):
    for i, row in enumerate(sheet1.rows):
        if i == 0:
            columns = {}
            for cell in row:
                columns[cell.value] = cell.col_idx - 1
                columns[cell.col_idx - 1] = cell.value
        if i > 0:
            id = row[columns[field]].value
            record = {columns[i]:x.value for i,x in enumerate(row)}
            if id in map:
                map[id]['a'].append(record)
            else:
                map[id] = {'a':[record]}

    for i, row in enumerate(sheet2.rows):
        if i == 0:
            columns = {}
            for cell in row:
                columns[cell.value] = cell.col_idx - 1
                columns[cell.col_idx - 1] = cell.value
        if i > 0:
            id = row[columns[field]].value
            record = {columns[i]:x.value for i,x in enumerate(row)}
            if id in map:
                if 'b' in map[id]:
                    map[id]['b'].append(record)
                else:
                    map[id]['b'] = [record]
            else:
                map[id] = {'b':[record]}
    return map


def compare_images(archive, f1, f2, args, dev, pages={}, content={}):
    assets = get_content({}, 'CMS_Content__c', f1['Assets'], f2['Assets'])
    contents = get_content({}, 'Slug__c', f1['Contents'], f2['Contents'])
    images = get_content({}, 'ContentDocumentId', f1['Files'], f2['Files'])
    merge = {}
    data = {}

    for c in contents.values():
        for row in assets.values():
            if 'a' in row:
                for asset in row['a']:
                    if 'a' in c and asset['CMS_Content__c'] == c['a'][0]['Id']:
                        slug = c['a'][0]['Slug__c']
                        key = asset['Asset_Type__c'] + '-' + asset['Name']
                        if slug in merge:
                            if 'a' in merge[slug]:
                                merge[slug]['a'][key] = asset
                            else:
                                merge[slug]['a'] = {key:asset}
                        else:
                            merge[slug] = {'a':{key:asset}}
            if 'b' in row:
                for asset in row['b']:
                    if 'b' in c and asset['CMS_Content__c'] == c['b'][0]['Id']:
                        slug = c['b'][0]['Slug__c']
                        key = asset['Asset_Type__c'] + '-' + asset['Name']
                        if slug in merge:
                            if 'b' in merge[slug]:
                                merge[slug]['b'][key] = asset
                            else:
                                merge[slug]['b'] = {key:asset}
                        else:
                            merge[slug] = {'b':{key:asset}}

    for slug, record in merge.items():
        if 'a' not in record:
            for b in record['b']:
                new = record['b'][b].copy()
                document_id = new['ContentDocument__c']
                b_file = get_image_by_id(document_id, images)
                a_file = get_image_by_title(b_file['Title'], images, 'a')
                b_content = get_content_by_id(new['CMS_Content__c'], content)
                a_content = get_content_by_slug(b_content['Slug__c'], content, 'a')
                if a_content:
                    if not a_file:
                        b64encoded = archive.read('content/' + document_id).decode('utf-8')
                        a_file = {'title' : b_file['Title'],'PathOnClient' : b_file['PathOnClient'],'VersionData' : b64encoded, 'Description': 'cms-asset'}
                        rs = dev.ContentVersion.create(a_file)
                        a_file['Id'] = rs['id']
                        db_file = dev.query('select ContentDocumentId from contentVersion where id = \'%s\''%(a_file['Id']))['records']
                        a_file['ContentDocumentId'] = db_file[0]['ContentDocumentId']

                    new.pop('Id')
                    new['CMS_Content__c'] = a_content['Id']
                    new['ContentDocument__c'] = a_file['ContentDocumentId']

                    dev.CMS_Asset__c.create(new)
                print(slug, b, 'remote not in dx')
            continue
        if 'b' not in record:
            for a in record['a']:
                print(slug, 'record not in remote', a)
            continue
        if 'a' and 'b' in record:
            assets = {}
            for b in record['b']:
                match = False
                if record['b'][b]['Asset_Type__c'] in assets:
                    assets[record['b'][b]['Asset_Type__c']]['b'] = record['b'][b]
                else:
                    assets[record['b'][b]['Asset_Type__c']] = {'b': record['b'][b]}
                for a in record['a']:
                    if record['a'][a]['Asset_Type__c'] in assets:
                        assets[record['a'][a]['Asset_Type__c']]['a'] = record['a'][a]
                    else:
                        assets[record['a'][a]['Asset_Type__c']] = {'a': record['a'][a]}
                    if a == b:
                        match = True
                if not match:
                    new = record['b'][b]
                    document_id = new['ContentDocument__c']
                    b_file = get_image_by_id(document_id, images)
                    a_file = get_image_by_title(b_file['Title'], images, 'a')
                    b_content = get_content_by_id(new['CMS_Content__c'], content)
                    a_content = get_content_by_slug(b_content['Slug__c'], content, 'a')
                    if a_content:
                        if not a_file:
                            b64encoded = archive.read('content/' + document_id).decode('utf-8')
                            a_file = {'title' : b_file['Title'],'PathOnClient' : b_file['PathOnClient'],'VersionData' : b64encoded, 'Description': 'cms-asset'}
                            rs = dev.ContentVersion.create(a_file)
                            a_file['Id'] = rs['id']
                            db_file = dev.query('select ContentDocumentId from contentVersion where id = \'%s\''%(a_file['Id']))['records']
                            a_file['ContentDocumentId'] = db_file[0]['ContentDocumentId']

                        new.pop('Id')
                        new['CMS_Content__c'] = a_content['Id']
                        new['ContentDocument__c'] = a_file['ContentDocumentId']

                    dev.CMS_Asset__c.update(assets[new['Asset_Type__c']]['a']['Id'], new)


def get_content_by_id(id, content):
    for slug in content:
        if 'a' in content[slug] and id == content[slug]['a']['Id']:
            return content[slug]['a']
        if 'b' in content[slug] and id == content[slug]['b']['Id']:
            return content[slug]['b']
    return None

def get_content_by_slug(slug, contents, side):
    for content in contents.values():
        if side in content:
            if side in content and content[side]['Slug__c'] == slug:
                return content[side]
    return None

def get_image_by_id(id, images):
    for iid in images:
        if id == iid:
            if 'a' in images[iid]:
                return images[iid]['a'][0]
            if 'b' in images[iid]:
                return images[iid]['b'][0]
    return None

def get_image_by_title(name, imageset, side):
    for images in imageset.values():
        if side in images:
            for image in images[side]:
                if image['Title'] == name:
                    return image
    return None





dev = Salesforce(username=args.username, password=args.password, sandbox=args.sandbox, security_token=args.token)

f1 = load_workbook(get_x(args.primary))
f2 = load_workbook(get_x(args.secondary))


pages = compare_tab('Pages', 'CMS_Page__c', 'Slug__c', f1, f2, args, dev)
content = compare_tab('Contents','CMS_Content__c', 'Slug__c', f1, f2, args, dev, pages=pages)
compare_images(zipfile.ZipFile(args.secondary, 'r'), f1, f2, args, dev, pages=pages, content=content)



